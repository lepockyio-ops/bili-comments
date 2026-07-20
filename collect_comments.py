#!/usr/bin/env python3
"""
BiliComments v2 — B 站视频评论收集器
=====================================
输入 B 站视频链接或 bvid → 抓取评论 → 输出增强的 XLSX
包含：
  * 原文评论
  * AI 日语翻译（Google Translate + Vocaloid 术语表预处理）
  * 意图自动分类（求授权/求歌词/情感深度/负面质疑等）
  * 推荐日语回复（3 条模板供 UP 挑选）
  * GOOGLETRANSLATE 公式列作为备用

用法：
    python collect_comments.py BV1FDNJ6BE4j
    python collect_comments.py BVxxx --max-pages 1               # 只抓 top 20
    python collect_comments.py BVxxx --include-replies           # 含楼中楼
    python collect_comments.py BVxxx --no-translate --no-reply   # 关闭 v2 特性

登录（可选，抓全量必须）：
    在 .env 里设置 BILI_SESSDATA=你的cookie 值
    或者环境变量 BILI_SESSDATA

依赖：httpx, openpyxl, pyyaml, deep-translator
"""
from __future__ import annotations

import argparse
import hashlib
import os
import random
import re
import sys
import time
import urllib.parse
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    import httpx
except ImportError:
    print("缺少 httpx。pip install httpx openpyxl pyyaml deep-translator", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("缺少 openpyxl。pip install httpx openpyxl pyyaml deep-translator", file=sys.stderr)
    sys.exit(1)


UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
CST = timezone(timedelta(hours=8))
SCRIPT_DIR = Path(__file__).parent.resolve()
DEFAULT_OUT_DIR = SCRIPT_DIR / "data"
ENV_FILE = SCRIPT_DIR / ".env"

MIXIN_KEY_ENC_TAB = [
    46, 47, 18, 2, 53, 8, 23, 32, 15, 50, 10, 31, 58, 3, 45, 35,
    27, 43, 5, 49, 33, 9, 42, 19, 29, 28, 14, 39, 12, 38, 41, 13,
    37, 48, 7, 16, 24, 55, 40, 61, 26, 17, 0, 1, 60, 51, 30, 4,
    22, 25, 54, 21, 56, 59, 6, 63, 57, 62, 11, 36, 20, 34, 44, 52,
]


# ============================================================================
# 工具
# ============================================================================
def parse_bvid(raw: str) -> str:
    raw = raw.strip()
    m = re.search(r"BV[0-9A-Za-z]{10}", raw)
    if m:
        return m.group(0)
    raise ValueError(f"无法从 '{raw}' 提取 bvid")


def fmt_time(ts: int) -> str:
    if not ts:
        return ""
    return datetime.fromtimestamp(ts, CST).strftime("%Y-%m-%d %H:%M")


def clean_msg(msg: str) -> str:
    if not msg:
        return ""
    return re.sub(r"\s+", " ", msg).strip()


def load_env_sessdata() -> str | None:
    val = os.environ.get("BILI_SESSDATA", "").strip()
    if val:
        return val
    if ENV_FILE.exists():
        with ENV_FILE.open(encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line.startswith("BILI_SESSDATA="):
                    return line.split("=", 1)[1].strip().strip('"').strip("'")
    return None


def _mixin_key(orig: str) -> str:
    return "".join(orig[i] for i in MIXIN_KEY_ENC_TAB)[:32]


def wbi_sign(params: dict, img_key: str, sub_key: str) -> dict:
    mk = _mixin_key(img_key + sub_key)
    params = dict(params)
    params["wts"] = int(time.time())
    sp = dict(sorted(params.items()))
    for k, v in sp.items():
        sp[k] = "".join(c for c in str(v) if c not in "!'()*")
    q = urllib.parse.urlencode(sp)
    sp["w_rid"] = hashlib.md5((q + mk).encode()).hexdigest()
    return sp


# ============================================================================
# B 站客户端
# ============================================================================
class BiliClient:
    def __init__(self, sessdata: str | None = None):
        self.sessdata = sessdata
        headers = {"User-Agent": UA, "Referer": "https://www.bilibili.com/"}
        cookies = {}
        if sessdata:
            cookies["SESSDATA"] = sessdata
        self.client = httpx.Client(
            timeout=20.0, headers=headers, cookies=cookies, follow_redirects=True,
        )
        self._bootstrapped = False
        self._img_key: str | None = None
        self._sub_key: str | None = None

    def _bootstrap(self):
        if self._bootstrapped:
            return
        try:
            self.client.get("https://www.bilibili.com/")
        except Exception:
            pass
        self._bootstrapped = True

    def _get_wbi_keys(self):
        if self._img_key:
            return
        self._bootstrap()
        r = self.client.get("https://api.bilibili.com/x/web-interface/nav")
        d = r.json()["data"]["wbi_img"]
        self._img_key = d["img_url"].rsplit("/", 1)[1].split(".")[0]
        self._sub_key = d["sub_url"].rsplit("/", 1)[1].split(".")[0]

    def close(self):
        self.client.close()

    def get_view(self, bvid: str) -> dict:
        self._bootstrap()
        r = self.client.get(
            "https://api.bilibili.com/x/web-interface/view",
            params={"bvid": bvid},
        )
        return r.json()

    def get_comments_page(self, oid: int, bvid: str, next_cursor: int = 0, mode: int = 3) -> dict:
        self._get_wbi_keys()
        params = {
            "next": next_cursor, "type": 1, "oid": oid, "mode": mode,
            "plat": 1, "web_location": 1315875, "seek_rpid": "",
        }
        signed = wbi_sign(params, self._img_key, self._sub_key)
        for attempt in range(3):
            try:
                r = self.client.get(
                    "https://api.bilibili.com/x/v2/reply/wbi/main",
                    params=signed,
                    headers={"Referer": f"https://www.bilibili.com/video/{bvid}"},
                )
                if "json" not in r.headers.get("content-type", "").lower():
                    time.sleep((attempt + 1) * 10)
                    continue
                return r.json()
            except httpx.HTTPError as e:
                print(f"  ! 网络错误 (cursor={next_cursor}): {e}", file=sys.stderr)
                time.sleep((attempt + 1) * 3)
        return {"code": -1, "message": "重试后仍失败"}

    def get_sub_replies_page(self, oid: int, root_rpid: int, pn: int, ps: int = 20) -> dict:
        for attempt in range(3):
            try:
                r = self.client.get(
                    "https://api.bilibili.com/x/v2/reply/reply",
                    params={"type": 1, "oid": oid, "root": root_rpid, "pn": pn, "ps": ps},
                )
                if "json" not in r.headers.get("content-type", "").lower():
                    time.sleep((attempt + 1) * 10)
                    continue
                return r.json()
            except httpx.HTTPError as e:
                print(f"  ! 子评论网络错误: {e}", file=sys.stderr)
                time.sleep((attempt + 1) * 3)
        return {"code": -1, "message": "重试后仍失败"}


# ============================================================================
# 数据提取
# ============================================================================
def extract_comment(c: dict, up_mid: int, is_sub: bool = False, parent_rpid: int = 0) -> dict:
    m = c.get("member") or {}
    level = ((m.get("level_info") or {}).get("current_level")) or 0
    vip = ((m.get("vip") or {}).get("vipStatus")) or 0
    nameplate = ((m.get("nameplate") or {}).get("name")) or ""
    mid = int(m.get("mid") or 0)

    return {
        "rpid": c.get("rpid"),
        "parent_rpid": parent_rpid,
        "is_sub": is_sub,
        "mid": mid,
        "uname": m.get("uname", ""),
        "level": level,
        "vip": "是" if vip == 1 else "",
        "is_up": "是" if (mid == up_mid and up_mid > 0) else "",
        "nameplate": nameplate,
        "sign": (m.get("sign") or "")[:80],
        "message": clean_msg((c.get("content") or {}).get("message", "")),
        "like": c.get("like", 0),
        "rcount": c.get("rcount", 0),
        "ctime": c.get("ctime", 0),
    }


# ============================================================================
# 采集
# ============================================================================
def collect_all(
    bvid: str,
    sort: str = "hot",
    max_pages: int | None = None,
    include_replies: bool = False,
) -> tuple[dict, list[dict]]:
    sessdata = load_env_sessdata()
    client = BiliClient(sessdata=sessdata)
    mode = {"hot": 3, "new": 2, "up-only": 1}.get(sort, 3)

    if sessdata:
        print(f"→ 使用 SESSDATA 登录（可抓全量评论）")
    else:
        print(f"⚠️  未提供 SESSDATA cookie — 只能拿到少数热门评论")

    print(f"→ 查询视频信息 {bvid}")
    view = client.get_view(bvid)
    if view.get("code") != 0:
        raise RuntimeError(f"视频信息查询失败: {view.get('message')}")

    v = view["data"]
    aid = v["aid"]
    up_mid = int((v.get("owner") or {}).get("mid") or 0)
    meta = {
        "bvid": bvid, "aid": aid, "title": v["title"],
        "up_name": (v.get("owner") or {}).get("name", ""),
        "up_mid": up_mid, "pubdate": v.get("pubdate"),
        "cover": v.get("pic", ""), "duration": v.get("duration"),
        "stat": v.get("stat", {}), "collected_at": int(time.time()),
        "logged_in": bool(sessdata),
    }

    print(f"→ 《{meta['title']}》 · UP {meta['up_name']}")
    print(f"→ 评论总数 (声称): {v['stat']['reply']}")
    print(f"→ 开始拉取评论 (sort={sort}, mode={mode})")

    comments: list[dict] = []
    seen_rpids: set[int] = set()
    next_cursor = 0
    page_idx = 0
    first_page = True

    while True:
        if max_pages and page_idx >= max_pages:
            print(f"  · 达到 max-pages={max_pages}，停止")
            break
        resp = client.get_comments_page(aid, bvid, next_cursor=next_cursor, mode=mode)
        if resp.get("code") != 0:
            print(f"  ! cursor={next_cursor} 失败: {resp.get('message')}")
            break
        data = resp.get("data") or {}
        replies = data.get("replies") or []
        cursor = data.get("cursor") or {}

        if first_page:
            top_replies = data.get("top_replies") or []
            upper = (data.get("upper") or {}).get("top")
            if upper:
                top_replies = [upper] + top_replies
            for c in top_replies:
                if c and c.get("rpid") and c["rpid"] not in seen_rpids:
                    seen_rpids.add(c["rpid"])
                    comments.append(extract_comment(c, up_mid))
            first_page = False

        new_count = 0
        for c in replies:
            if c.get("rpid") in seen_rpids:
                continue
            seen_rpids.add(c["rpid"])
            comments.append(extract_comment(c, up_mid))
            new_count += 1

        page_idx += 1
        print(f"  · page {page_idx} (cursor={next_cursor}) +{new_count} (累计 {len(comments)})"
              f" · is_end={cursor.get('is_end')} · all_count={cursor.get('all_count')}")

        if cursor.get("is_end") or not replies:
            break
        next_cursor = cursor.get("next", 0)
        if not next_cursor:
            break
        time.sleep(random.uniform(1.5, 3.0))

    if include_replies:
        main_with_replies = [c for c in comments if c["rcount"] > 0 and not c["is_sub"]]
        print(f"\n→ 开始拉取楼中楼（{len(main_with_replies)} 条主评论有回复）")
        for i, main in enumerate(main_with_replies, 1):
            root = main["rpid"]
            sub_pn = 1
            sub_pages = 0
            while True:
                resp = client.get_sub_replies_page(aid, root, pn=sub_pn)
                if resp.get("code") != 0:
                    break
                subs = (resp.get("data") or {}).get("replies") or []
                if not subs:
                    break
                for c in subs:
                    if c.get("rpid") in seen_rpids:
                        continue
                    seen_rpids.add(c["rpid"])
                    comments.append(extract_comment(c, up_mid, is_sub=True, parent_rpid=root))
                sub_pages += 1
                sub_pn += 1
                time.sleep(random.uniform(0.8, 2.0))
                if sub_pages >= 10:
                    break
            if i % 5 == 0:
                print(f"  · 子评论进度 {i}/{len(main_with_replies)}")

    client.close()
    main_count = sum(1 for c in comments if not c["is_sub"])
    sub_count = sum(1 for c in comments if c["is_sub"])
    print(f"\n✓ 采集完成：{len(comments)} 条评论（主 {main_count} + 楼中楼 {sub_count}）")
    return meta, comments


# ============================================================================
# v2 增强：翻译 + 回复推荐
# ============================================================================
def enrich_comments(
    comments: list[dict],
    do_translate: bool = True,
    do_reply: bool = True,
    target_lang: str = "ja",
) -> None:
    """给评论添加 ai_translation / intent / replies 三个字段（in-place）"""
    if do_translate:
        try:
            from translator import translate_comments as _translate_batch
            messages = [c["message"] for c in comments]
            translations = _translate_batch(messages, target_lang=target_lang, verbose=True)
            for c, t in zip(comments, translations):
                c["ai_translation"] = t or ""
        except Exception as e:
            print(f"⚠️  翻译失败（将 fallback 到公式列）: {e}")
            for c in comments:
                c["ai_translation"] = ""
    else:
        for c in comments:
            c["ai_translation"] = ""

    if do_reply:
        try:
            from reply_gen import generate_replies_for_comments as _gen_replies
            results = _gen_replies(comments, k=3)
            for c, (intent, replies) in zip(comments, results):
                c["intent"] = intent
                c["reply_1"] = replies[0] if len(replies) > 0 else ""
                c["reply_2"] = replies[1] if len(replies) > 1 else ""
                c["reply_3"] = replies[2] if len(replies) > 2 else ""
            print(f"✓ 已为 {len(comments)} 条评论生成意图标签与回复建议")
        except Exception as e:
            print(f"⚠️  回复生成失败: {e}")
            for c in comments:
                c["intent"] = ""
                c["reply_1"] = c["reply_2"] = c["reply_3"] = ""
    else:
        for c in comments:
            c["intent"] = ""
            c["reply_1"] = c["reply_2"] = c["reply_3"] = ""


# ============================================================================
# XLSX 输出（v2 加宽版）
# ============================================================================
def export_xlsx(meta: dict, comments: list[dict], out_path: Path, target_lang: str = "ja"):
    """
    v3.0 简化版：
    - 删除 Sheet 1 「评论」（用户不要全量评论）
    - Sheet 1 变成「Top 20 高赞」
    - Sheet 2「视频信息」精简：去掉登录状态/主评论数/楼中楼数/合计/UP回复数/大会员数/平均等级/最高点赞
    """
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    v2_fill = PatternFill("solid", fgColor="2E7D32")
    up_fill = PatternFill("solid", fgColor="FFF3CD")

    # ---------- Sheet 1: Top 20 高赞（v3.0 现在是主表）----------
    ws = wb.active
    ws.title = "Top 20 高赞"
    ws.append(["排名", "用户名", "点赞", "评论中文", "AI日语翻译", "链接"])
    for col_idx in range(1, 7):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = v2_fill if col_idx == 5 else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    top20 = sorted(
        [c for c in comments if not c["is_sub"]],
        key=lambda c: c["like"], reverse=True,
    )[:20]
    for i, c in enumerate(top20, start=2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=c["uname"])
        ws.cell(row=i, column=3, value=c["like"])
        ws.cell(row=i, column=4, value=c["message"])
        ws.cell(row=i, column=5, value=c.get("ai_translation", ""))
        ws.cell(row=i, column=6, value=f"https://www.bilibili.com/video/{meta['bvid']}#reply{c['rpid']}")

    widths = [6, 22, 8, 60, 60, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # UP 主评论行自动黄色高亮
    for i, c in enumerate(top20, start=2):
        if c.get("is_up") == "是":
            for col_idx in range(1, 7):
                ws.cell(row=i, column=col_idx).fill = up_fill

    # ---------- Sheet 2: 视频信息（v3.0 精简版）----------
    ws2 = wb.create_sheet("视频信息")
    stat = meta.get("stat") or {}
    info_rows = [
        ("视频标题", meta["title"]),
        ("BV 号", meta["bvid"]),
        ("AV 号", meta["aid"]),
        ("UP 主", meta["up_name"]),
        ("UP UID", meta["up_mid"]),
        ("发布时间", fmt_time(meta.get("pubdate", 0))),
        ("时长（秒）", meta.get("duration", 0)),
        ("视频链接", f"https://www.bilibili.com/video/{meta['bvid']}"),
        ("", ""),
        ("=== 视频数据 ===", ""),
        ("播放", stat.get("view", 0)),
        ("点赞", stat.get("like", 0)),
        ("投币", stat.get("coin", 0)),
        ("收藏", stat.get("favorite", 0)),
        ("分享", stat.get("share", 0)),
        ("评论（声称）", stat.get("reply", 0)),
        ("弹幕", stat.get("danmaku", 0)),
        ("", ""),
        ("采集时间", fmt_time(meta.get("collected_at", 0))),
    ]
    for r in info_rows:
        ws2.append(r)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 60
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"\n✓ XLSX 已保存：{out_path.resolve()}")
    print(f"  → v3.0 精简版：Top 20 高赞（主表）+ 视频信息（副表）")
    print(f"  → UP 主评论行自动黄色高亮")


# ============================================================================
# 入口
# ============================================================================
def main():
    p = argparse.ArgumentParser(description="BiliComments v2 — 评论收集 + AI 翻译 + 回复推荐")
    p.add_argument("bvid", help="BVxxx 或视频链接")
    p.add_argument("--out", help="输出 XLSX 路径（默认 data/<bvid>.xlsx）")
    p.add_argument("--sort", choices=["hot", "new", "up-only"], default="hot")
    p.add_argument("--max-pages", type=int, help="最多拉取多少页评论")
    p.add_argument("--include-replies", action="store_true", help="同时拉楼中楼")
    p.add_argument("--target-lang", default="ja", help="翻译目标语言（默认 ja 日语）")
    p.add_argument("--no-translate", action="store_true", help="关闭 v2 AI 翻译（只保留公式列）")
    p.add_argument("--no-reply", action="store_true", help="关闭 v2 回复推荐")
    args = p.parse_args()

    bvid = parse_bvid(args.bvid)
    meta, comments = collect_all(
        bvid, sort=args.sort,
        max_pages=args.max_pages, include_replies=args.include_replies,
    )

    if comments:
        print()
        enrich_comments(
            comments,
            do_translate=not args.no_translate,
            do_reply=not args.no_reply,
            target_lang=args.target_lang,
        )

    out_path = Path(args.out) if args.out else (DEFAULT_OUT_DIR / f"{bvid}.xlsx")
    export_xlsx(meta, comments, out_path, target_lang=args.target_lang)


if __name__ == "__main__":
    main()
